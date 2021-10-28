from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from PySide6 import QtCore
from PySide6.QtWidgets import QApplication
import pyperclip
import pandas as pd
import time
import os
import sys

class Web_HitHorizon(QtCore.QThread):

    #This is the signal that will be emitted during the processing.
    #By including int as an argument, it lets the signal know to expect
    #an integer argument when emitting.
    updateProgress = QtCore.Signal(int)


        

    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False, 
                           **to_excel_kwargs):

        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name, 
                startrow=startrow if startrow is not None else 0, 
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()


    # get folder path which has been converted into binary file. (sys._MEIPASS)
    @staticmethod
    def resource_path(relative_path):
        try:
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.dirname(__file__)
        return os.path.join(base_path, relative_path)


    def create_driver(self):
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("prefs", {"intl.accept_languages": "en-US"})
        chrome_options.add_experimental_option("excludeSwitches", ['enable-logging'])
        return webdriver.Chrome(executable_path=self.resource_path('ChromeDriver\chromedriver.exe'), options=chrome_options)
        # return webdriver.Chrome(executable_path = os.path.relpath('ChromeDriver\chromedriver.exe'), options=chrome_options)


    # df_company = pd.read_excel(r"C:\Users\ADMIN\Documents\Aufinia\Data_Test.xlsx",sheet_name='Euro', converters={'LFA1_LIFNR':str,'LFA1_STCD1':str},usecols={'LFA1_LIFNR','LFA1_NAME1','COUNTRY','LFA1_STCD1','LFA1_LAND1'})

    
    def crawling_execution(self,df_company):

        df_company_output = df_company[['Supplier_number','Supplier_name','Country','Tax_code','Country_code']]
        
        driver = self.create_driver()
        driver.get('https://www.hithorizons.com/search?Name=')

        try:
            for idx in range(len(df_company_output)):

                time.sleep(3)
                # get company_name, country_name from dataframe
                company_name = df_company_output.loc[idx,'Supplier_name']
                country_name = df_company_output.loc[idx,'Country']
        
                
        
                # send company_name keys
                driver.implicitly_wait(15)
                company_name_input = driver.find_element(By.ID, "Name")
                company_name_input.clear()
                company_name_input.send_keys(company_name)
        
        
                # send country_name keys
                driver.implicitly_wait(15)
                country_name_input = driver.find_element(By.ID, "Address")
                country_name_input.clear()
                country_name_input.send_keys(country_name)
        
        
                #click search button
                search_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='main-content']/div[1]/div/form/div[1]/div[3]/button")))
                driver.execute_script("arguments[0].click();", search_button)
        
                time.sleep(3)
                
                try:
                
                    driver.find_element(By.XPATH, "//*[@id='main-content']/div[2]/div[1]/div/div[1]/div[1]/div[1]/h3/a")
                
                except NoSuchElementException:
                
                    df_company_output.loc[idx,'Exception'] = "No information found"
                    
                    # emit signal to main_GUI
                    QApplication.processEvents()
                    self.updateProgress.emit(((idx+1)  * 100)/len(df_company_output))
                    time.sleep(0.1)
                    
                    print(str(idx) +": "+str(company_name)+ "  not found") 
        
                else:
                
                    print(str(idx) +": "+str(company_name)) 
        
                    driver.refresh()
                    #click first href link
                    first_result = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='main-content']/div[2]/div[1]/div/div[1]/div[1]/div[1]/h3/a")))
                    first_result.click()
        
                    driver.implicitly_wait(10)
                    # click copy clipboard
                    clipboard_button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[@id='chart-content']/div/div/div[1]/div[1]/div[2]/div/div[1]/div")))
                    clipboard_button.click()
                    clipboard_text= pyperclip.paste()
        
        
                    text_line = clipboard_text.splitlines()
        
                    df_company_output.loc[idx,'Name found in HitHorizons'] = text_line[0].strip()
                    df_company_output.loc[idx,'Address found in HitHorizons'] = text_line[1].strip()
        
                    if(len(text_line) == 3):
                        df_company_output.loc[idx,'Tax code found in HitHorizons'] = text_line[2].strip()
                    else:
                        df_company_output.loc[idx,'Tax code found in HitHorizons'] = 'None'
        
                    df_company_output.loc[idx,'Industry found in HitHorizons'] = driver.find_element(By.XPATH,  "//*[@id='chart-content']/div/div/div[1]/div[1]/div[2]/div/div[2]/ul/li[1]/span").text.strip()
                    df_company_output.loc[idx,'URL of HitHorizons'] = driver.current_url
        
                    # emit signal to main_GUI
                    QApplication.processEvents()
                    self.updateProgress.emit(((idx+1)  * 100)/len(df_company_output))
                    time.sleep(0.1)
                    
                    # Get back to previous page
                    driver.back()

            driver.close()
            df_company_output = df_company_output.fillna('None')
            return df_company_output
            

        except WebDriverException:

            df_company_output = df_company_output.fillna('None')
            return df_company_output
        




    
   


